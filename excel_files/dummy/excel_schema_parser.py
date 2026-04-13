"""
excel_schema_parser.py
-----------------------
Parses an ETL mapping-spec Excel file and produces source/target schemas,
column mapping rules, and join definitions.

Special sheets:
  _joins  — defines cross-table joins for enriched extracts (see README)

Output JSON structure:
{
  "joins": {
    "<sheet_name>": [
      {
        "join_alias":    "sup",
        "join_table":    "supplier_master",
        "join_db":       "",          // blank = same DB
        "join_type":     "LEFT",
        "left_key":      "supplier_id",
        "right_key":     "supplier_id",
        "fetch_columns": ["supplier_nm", "country_cd"]
      }, ...
    ]
  },
  "tables": {
    "<sheet_name>": {
      "source_table":   "...",
      "target_table":   "...",
      "joins":          [...],        // resolved from _joins sheet
      "source_schema":  { "table": ..., "primary_keys": [...], "columns": [...] },
      "target_schema":  { "table": ..., "primary_keys": [...], "columns": [...] },
      "column_mappings": [...]
    }
  }
}

Transform-rule syntax for joined columns:
  <alias>.<column>               e.g.  sup.supplier_nm
  UPPER(<alias>.<column>)        e.g.  UPPER(sup.country_cd)
  <alias>.<col1> + ' ' + <alias>.<col2>
  Any expression mixing main-table cols and alias.col references is valid.

Usage:
    python excel_schema_parser.py mapping_spec.xlsx [--output schemas.json] [--pretty]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

import openpyxl

# ── Constants ──────────────────────────────────────────────────────────────

VALID_TRANSFORM_TYPES = {"direct", "rename", "cast", "derived", "drop", "constant"}
JOINS_SHEET           = "_joins"

HEADER_ALIASES: dict[str, str] = {
    "src_col_name":    "src_col_name",
    "src_dtype":       "src_dtype",
    "src_nullable":    "src_nullable",
    "src_is_pk":       "src_is_pk",
    "src_is_unique":   "src_is_unique",
    "transform_type":  "transform_type",
    "transform_rule":  "transform_rule",
    "tgt_col_name":    "tgt_col_name",
    "tgt_dtype":       "tgt_dtype",
    "tgt_nullable":    "tgt_nullable",
    "tgt_default_val": "tgt_default_val",
    "force_dtype":     "force_dtype",
    "null_handling":   "null_handling",
    "notes":           "notes",
}

# _joins sheet column names (row 1 = headers, no section rows)
JOINS_HEADERS = {
    "sheet_name", "join_alias", "join_table", "join_db",
    "join_type", "left_key", "right_key", "fetch_columns",
}


# ── Data models ────────────────────────────────────────────────────────────

@dataclass
class JoinDef:
    join_alias:    str
    join_table:    str
    join_db:       str          # blank = same DB as source
    join_type:     str          # LEFT / INNER / RIGHT
    left_key:      str          # column in main source table
    right_key:     str          # column in join table
    fetch_columns: list[str]    # columns to pull from join table

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class SourceColumn:
    name: str
    dtype: str
    nullable: bool
    is_pk: bool
    is_unique: bool

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class TargetColumn:
    name: str
    dtype: str
    nullable: bool
    default_val: Any
    effective_dtype: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ColumnMapping:
    source_columns: list[str]       # main-table columns referenced
    join_columns:   list[str]       # "alias.col" references from join tables
    transform_type: str
    transform_rule: str | None
    target_column:  str | None
    force_dtype:    str | None
    null_handling:  str | None
    notes:          str | None

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class TableSpec:
    sheet_name:     str
    source_table:   str
    target_table:   str
    joins:          list[dict] = field(default_factory=list)
    source_schema:  dict       = field(default_factory=dict)
    target_schema:  dict       = field(default_factory=dict)
    column_mappings: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "source_table":    self.source_table,
            "target_table":    self.target_table,
            "joins":           self.joins,
            "source_schema":   self.source_schema,
            "target_schema":   self.target_schema,
            "column_mappings": self.column_mappings,
        }


# ── Helpers ────────────────────────────────────────────────────────────────

def _cell_val(cell) -> str | None:
    v = cell.value
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _bool_flag(val: str | None) -> bool:
    if val is None:
        return False
    return val.upper() in {"Y", "YES", "TRUE", "1"}


def _parse_name_row(ws) -> tuple[str, str]:
    """Row 1: '<source> → <target>'  (accepts → or ->)."""
    raw = _cell_val(ws.cell(row=1, column=1))
    for sep in ("→", "->"):
        if raw and sep in raw:
            parts = raw.split(sep, 1)
            return parts[0].strip(), parts[1].strip()
    raise ValueError(
        f"Row 1 of sheet '{ws.title}' must be '<source> -> <target>'. Got: {raw!r}"
    )


def _parse_subheaders(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = _cell_val(ws.cell(row=3, column=col_idx))
        if raw:
            key = raw.lower().strip()
            canonical = HEADER_ALIASES.get(key, key)
            mapping[canonical] = col_idx
    return mapping


def _row_as_dict(ws, row_idx: int, header_map: dict[str, int]) -> dict[str, str | None]:
    return {
        f: _cell_val(ws.cell(row=row_idx, column=col_idx))
        for f, col_idx in header_map.items()
    }


def _extract_join_refs(rule: str) -> list[str]:
    """Return all 'alias.column' tokens found in a transform rule."""
    return re.findall(r'\b([a-z_][a-z0-9_]*)\.([a-z_][a-z0-9_]*)\b', rule, re.IGNORECASE)


# ── _joins sheet parser ────────────────────────────────────────────────────

def parse_joins_sheet(ws) -> dict[str, list[JoinDef]]:
    """
    Parse the _joins sheet.
    Row 1 = column headers (no section rows, no merged cells).
    Returns {sheet_name: [JoinDef, ...]}
    """
    # Read header row
    header_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = _cell_val(ws.cell(row=1, column=col_idx))
        if raw:
            header_map[raw.lower().strip()] = col_idx

    missing = JOINS_HEADERS - set(header_map.keys())
    if missing:
        raise ValueError(f"_joins sheet is missing columns: {missing}")

    result: dict[str, list[JoinDef]] = {}

    for row_idx in range(2, ws.max_row + 1):
        row = {k: _cell_val(ws.cell(row=row_idx, column=v)) for k, v in header_map.items()}

        if all(v is None for v in row.values()):
            break

        sheet_name   = row.get("sheet_name")
        join_alias   = row.get("join_alias")
        join_table   = row.get("join_table")
        join_db      = row.get("join_db") or ""
        join_type    = (row.get("join_type") or "LEFT").upper()
        left_key     = row.get("left_key")
        right_key    = row.get("right_key")
        fetch_raw    = row.get("fetch_columns") or ""

        if not all([sheet_name, join_alias, join_table, left_key, right_key]):
            print(f"  WARNING: _joins row {row_idx} incomplete, skipping.")
            continue

        fetch_columns = [c.strip() for c in fetch_raw.split(",") if c.strip()]

        jd = JoinDef(
            join_alias=join_alias,
            join_table=join_table,
            join_db=join_db,
            join_type=join_type,
            left_key=left_key,
            right_key=right_key,
            fetch_columns=fetch_columns,
        )
        result.setdefault(sheet_name, []).append(jd)

    return result


# ── Mapping sheet parser ───────────────────────────────────────────────────

def parse_sheet(ws, joins_for_sheet: list[JoinDef] | None = None) -> TableSpec:
    source_table, target_table = _parse_name_row(ws)
    header_map = _parse_subheaders(ws)

    required = {"src_col_name", "src_dtype", "transform_type"}
    missing = required - set(header_map.keys())
    if missing:
        raise ValueError(f"Sheet '{ws.title}': missing required sub-headers: {missing}")

    # Build alias → JoinDef index for rule validation
    alias_index: dict[str, JoinDef] = {}
    join_fetched_cols: set[str] = set()   # all columns available from joins
    for jd in (joins_for_sheet or []):
        alias_index[jd.join_alias.lower()] = jd
        for fc in jd.fetch_columns:
            join_fetched_cols.add(f"{jd.join_alias.lower()}.{fc.lower()}")

    source_cols:  list[SourceColumn] = []
    target_cols:  list[TargetColumn] = []
    mappings:     list[ColumnMapping] = []
    errors:       list[str] = []

    for row_idx in range(4, ws.max_row + 1):
        row = _row_as_dict(ws, row_idx, header_map)

        if all(v is None for v in row.values()):
            break

        src_name      = row.get("src_col_name")
        src_dtype     = row.get("src_dtype")
        src_nullable  = _bool_flag(row.get("src_nullable"))
        src_is_pk     = _bool_flag(row.get("src_is_pk"))
        src_is_unique = _bool_flag(row.get("src_is_unique"))

        transform_type = (row.get("transform_type") or "").lower().strip()
        transform_rule = row.get("transform_rule")

        tgt_name      = row.get("tgt_col_name")
        tgt_dtype     = row.get("tgt_dtype")
        tgt_nullable  = _bool_flag(row.get("tgt_nullable"))
        tgt_default   = row.get("tgt_default_val")

        force_dtype   = row.get("force_dtype")
        null_handling = row.get("null_handling")
        notes         = row.get("notes")

        # ── Validation ──────────────────────────────────────────────────────
        if not transform_type:
            errors.append(f"Row {row_idx}: transform_type is required.")
            continue

        if transform_type not in VALID_TRANSFORM_TYPES:
            errors.append(f"Row {row_idx}: unknown transform_type '{transform_type}'.")
            continue

        has_source = bool(src_name)
        has_target = bool(tgt_name)

        if transform_type == "drop":
            if not has_source:
                errors.append(f"Row {row_idx}: 'drop' requires src_col_name.")
                continue
        elif transform_type == "constant":
            if not has_target or not transform_rule:
                errors.append(f"Row {row_idx}: 'constant' requires tgt_col_name + transform_rule.")
                continue
        else:
            if not has_target:
                errors.append(f"Row {row_idx}: '{transform_type}' requires tgt_col_name.")
                continue
            if transform_type in {"cast", "derived"} and not transform_rule:
                errors.append(f"Row {row_idx}: '{transform_type}' requires transform_rule.")
                continue
            if not has_source and not transform_rule:
                errors.append(f"Row {row_idx}: no src_col_name and no transform_rule.")
                continue

        # ── Detect join references in transform_rule ────────────────────────
        join_col_refs: list[str] = []   # "alias.col" strings
        if transform_rule and alias_index:
            refs = _extract_join_refs(transform_rule)
            for alias, col in refs:
                a_lower = alias.lower()
                if a_lower in alias_index:
                    ref = f"{a_lower}.{col.lower()}"
                    join_col_refs.append(ref)
                    # Warn if column wasn't listed in fetch_columns
                    if ref not in join_fetched_cols:
                        print(
                            f"  WARNING: sheet '{ws.title}' row {row_idx}: "
                            f"rule references '{ref}' but '{col}' is not in "
                            f"fetch_columns for alias '{a_lower}'. "
                            f"Adding it automatically."
                        )
                        alias_index[a_lower].fetch_columns.append(col)
                        join_fetched_cols.add(ref)

        # ── Build source column ──────────────────────────────────────────────
        if has_source:
            existing = {c.name for c in source_cols}
            if src_name not in existing:
                source_cols.append(SourceColumn(
                    name=src_name,
                    dtype=src_dtype or "UNKNOWN",
                    nullable=src_nullable,
                    is_pk=src_is_pk,
                    is_unique=src_is_unique,
                ))

        # ── Build target column ──────────────────────────────────────────────
        if has_target:
            effective_dtype = force_dtype if force_dtype else (tgt_dtype or "UNKNOWN")
            existing = {c.name for c in target_cols}
            if tgt_name not in existing:
                target_cols.append(TargetColumn(
                    name=tgt_name,
                    dtype=tgt_dtype or "UNKNOWN",
                    nullable=tgt_nullable,
                    default_val=tgt_default,
                    effective_dtype=effective_dtype,
                ))

        # ── Build mapping ────────────────────────────────────────────────────
        if has_source and transform_type != "drop":
            source_columns = [src_name]
        elif transform_type == "drop":
            source_columns = [src_name] if has_source else []
        else:
            source_columns = []

        mappings.append(ColumnMapping(
            source_columns=source_columns,
            join_columns=join_col_refs,
            transform_type=transform_type,
            transform_rule=transform_rule,
            target_column=tgt_name,
            force_dtype=force_dtype,
            null_handling=null_handling,
            notes=notes,
        ).to_dict())

    if errors:
        raise ValueError(f"Sheet '{ws.title}' has {len(errors)} error(s):\n  " + "\n  ".join(errors))

    # ── Primary keys ─────────────────────────────────────────────────────────
    src_pks = [c.name for c in source_cols if c.is_pk]
    tgt_pks = []
    for m in mappings:
        if m["transform_type"] in {"direct", "rename", "cast"} and m["source_columns"]:
            sc = next((c for c in source_cols if c.name == m["source_columns"][0]), None)
            if sc and sc.is_pk and m["target_column"]:
                tgt_pks.append(m["target_column"])

    return TableSpec(
        sheet_name=ws.title,
        source_table=source_table,
        target_table=target_table,
        joins=[jd.to_dict() for jd in (joins_for_sheet or [])],
        source_schema={
            "table": source_table,
            "primary_keys": src_pks,
            "columns": [c.to_dict() for c in source_cols],
        },
        target_schema={
            "table": target_table,
            "primary_keys": tgt_pks,
            "columns": [c.to_dict() for c in target_cols],
        },
        column_mappings=mappings,
    )


# ── Top-level entry point ──────────────────────────────────────────────────

def parse_mapping_spec(excel_path: str | Path) -> dict:
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    # 1. Parse _joins sheet first (if present)
    joins_by_sheet: dict[str, list[JoinDef]] = {}
    if JOINS_SHEET in wb.sheetnames:
        print(f"  Found '{JOINS_SHEET}' sheet — parsing join definitions...")
        raw_joins = parse_joins_sheet(wb[JOINS_SHEET])
        joins_by_sheet = raw_joins
        print(f"  Join definitions loaded for: {list(raw_joins.keys())}")

    # 2. Parse mapping sheets
    result: dict[str, Any] = {"joins": {}, "tables": {}}

    for sheet_name in wb.sheetnames:
        if sheet_name == JOINS_SHEET:
            continue
        ws = wb[sheet_name]
        jds = joins_by_sheet.get(sheet_name, [])
        spec = parse_sheet(ws, jds)
        result["tables"][sheet_name] = spec.to_dict()
        if spec.joins:
            result["joins"][sheet_name] = spec.joins

    return result


# ── CLI ────────────────────────────────────────────────────────────────────

def main(excel_file, output, pretty):
    excel_path = Path(excel_file)
    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found.", file=sys.stderr)
        return 1

    try:
        result = parse_mapping_spec(excel_path)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    out = Path(output) if output else excel_path.with_name(excel_path.stem + "_schemas.json")
    out.write_text(json.dumps(result, indent=2 if pretty else None, default=str), encoding="utf-8")
    print(f"\nOK  Parsed {len(result['tables'])} table(s)  ->  {out}")

    for sheet, spec in result["tables"].items():
        n_joins = len(spec.get("joins", []))
        join_str = f"  [{n_joins} join(s)]" if n_joins else ""
        print(
            f"   [{sheet}]  {spec['source_table']} "
            f"({len(spec['source_schema']['columns'])} src cols) "
            f"-> {spec['target_table']} "
            f"({len(spec['target_schema']['columns'])} tgt cols)"
            f"{join_str}"
        )
    return 0


if __name__ == "__main__":
    excel_file = "vehicle_schema_mappings.xlsx"
    output = "vehicle_schemas.json"
    pretty = True

    sys.exit(main(excel_file, output, pretty))